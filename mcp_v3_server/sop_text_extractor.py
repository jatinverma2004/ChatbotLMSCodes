 
from pathlib import Path
import pdfplumber
from docx import Document
import csv
import json
from openpyxl import load_workbook
from PIL import Image
import io
 
def extract_text(file_path: str, file_name: str) -> str:
    """Extract text from various file formats"""
    
    ext = Path(file_path).suffix.lower()
    
    try:
        # PDF files
        if ext == '.pdf':
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages[:10]:  # First 10 pages
                    text += page.extract_text() or ""
            return text[:5000]
        
        # Word documents
        elif ext in ['.docx', '.doc']:
            try:
                doc = Document(file_path)
                text = "\n".join([para.text for para in doc.paragraphs])
                return text[:5000]
            except:
                return "[Could not extract text from DOCX]"
        
        # Excel files
        elif ext in ['.xlsx', '.xls']:
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                text = ""
                for row in list(ws.iter_rows(max_row=20)):
                    for cell in row:
                        if cell.value:
                            text += str(cell.value) + " | "
                    text += "\n"
                return text[:5000]
            except:
                return "[Could not extract text from Excel]"
        
        # CSV files
        elif ext == '.csv':
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                lines = list(reader)[:20]
                text = "\n".join([" | ".join(row) for row in lines])
            return text[:5000]
        
        # Text files
        elif ext == '.txt':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()[:5000]
        
        # Images (extract metadata and dimensions)
        elif ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
            try:
                img = Image.open(file_path)
                info = f"Image: {img.format} | Size: {img.size} | Mode: {img.mode}"
                if img.info:
                    info += f" | Info: {str(img.info)[:200]}"
                return info
            except:
                return "[Could not process image]"
        
        else:
            return "[File type not supported for text extraction]"
    
    except Exception as e:
        return f"[Error extracting text: {str(e)}]"
 
def get_file_preview(text: str, max_chars: int = 500) -> str:
    """Get preview of extracted text"""
    if not text:
        return ""
    return text[:max_chars] + ("..." if len(text) > max_chars else "")
 
